import openpyxl


class ExcelUtility:

    def get_data_from_excel(pageName):
        file_url = "data/userinfo.xlsx"
        file = openpyxl.load_workbook(file_url)
        page = file[pageName]
        line_number = page.max_row
        column_number = page.max_column
        data = []
        for i in range(2, line_number + 1):
            line = []
            for j in range(1, column_number + 1):
                line.append(page.cell(i, j).value)
            data.append(line)
        return data
